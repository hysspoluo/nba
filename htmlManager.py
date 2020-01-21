from bs4 import BeautifulSoup
import os
import time
from selenium import webdriver
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook

path = os.getcwd()
folderName = path+"\\download" #下载文件的路径
playoffgame = ["东部第一圈","西部第一圈","东部第二圈","西部第二圈","东部决赛","西部决赛","总决赛"]

'''
#数据下载
'''

def source2Html(Source,htmlName,season):
    # 目录不存在则创建目录
    isExists = os.path.exists(folderName+"\\"+season)
    if not isExists:
        os.makedirs(folderName+"\\"+season)
    fp = open(htmlName, "w", encoding='utf-8')
    fp.write(Source)
    fp.close()


#季前赛下载
def downloadPreSeason(driver,season):
    try:
        driver.find_element_by_id("menu3").click()
    except:
        print("没有季前赛")
        return
    htmlName = folderName+"\\"+season+"\\"+"季前赛"
    source2Html(driver.page_source,htmlName,season)

#常规赛下载
def downloadregularSeason(driver,season):
    try:
        driver.find_element_by_id("menu1").click()
    except:
        print("没有常规赛")
        return
    #获取月份标签
    tableElement = driver.find_element_by_id("yearmonthTable2")
    monthNum = len(tableElement.find_elements_by_class_name("lsm2"))
    num = 0
    while 1:
        if num == monthNum:
            break
        driver.find_element_by_id("yearmonthTable2").find_elements_by_class_name("lsm2")[num].click()
        num += 1
        time.sleep(0.5)
        htmlName = folderName+"\\"+season+"\\"+"常规赛_"+str(num)
        source2Html(driver.page_source, htmlName,season)

#季后赛下载
def downloadPlayOff(driver,season):
    try:
        driver.find_element_by_id("menu2").click()
    except:
        print("没有季后赛")
        return
    for game in playoffgame:
        str = "//*[text()='%s']" %(game)
        try:
            driver.find_element_by_xpath(str).click()
            time.sleep(0.5)
        except:
            print("季后赛还没开始")
            return
        htmlName = folderName+"\\"+season+"\\"+"季后赛_"+game
        source2Html(driver.page_source,htmlName,season)


def downloadNbaData(seasonLink):
    myDriver = webdriver.Chrome()  # 加载chrome内核
    # time.sleep(5)
    myDriver.get(seasonLink)  # 打开指定页面
    myDriver.maximize_window()  # 最大化屏幕
    #开始读取数据，先循环选择赛季
    # 获得赛季选项选项
    try:
        seasonSelect = Select(myDriver.find_element_by_name("seasonList"))
        seasonOptions = seasonSelect.options  # 获得select中的所有选项名称
    except:
        print("加载错误")
        return
    else:
        seasonName = []  # 赛季选项
        for team in seasonOptions:
            seasonName.append(team.text)
    #开始读取数据
        for item in seasonName:
            try:
                Select(myDriver.find_element_by_name("seasonList")).select_by_visible_text(item)
            except:
                print("加载赛季失败")
                return
            else:
                time.sleep(1)
                #开始进行数据分析
                #分析季前赛
                downloadPreSeason(myDriver,item)
                #分析常规赛
                downloadregularSeason(myDriver,item)
                # 分析季后赛
                downloadPlayOff(myDriver, item)

    myDriver.quit()

'''
#数据分析
'''
def analysHtml(htmlName):
    soup = BeautifulSoup(open(htmlName,"r",encoding='utf-8'),'html.parser')
    table = soup.find("table", id="scheTab")
    trs = table.find_all("tr",bgcolor = ["#E3EEF9","#ffffff"])#传递颜色参数list作为过滤条件
    Schedule_List = []
    for tr in trs:
        gameList = []
        for td in tr.find_all("td"):
            gameList.append(td.text)
        Schedule_List.append(gameList)
    #返回每页的比赛列表
    return Schedule_List


def gethtmlpath():
    sheetNames = {}
    dirs = os.listdir(folderName)
    for dir in dirs:
        files = os.listdir(folderName+"\\"+dir)
        sheetNames[dir]= files
    return sheetNames

def createxls(seasons):
    titles = ["类型", "时间", "主队", "比分", "客队", "让分", "总分", "资料", "半场"]
    num = 0
    isExists = os.path.exists(path + "\\gameData.xlsx")
    if not isExists:
        print("文件不存在")
        wb = Workbook()
        for season in seasons:
            wb.create_sheet(season)
            sheet = wb.get_sheet_by_name(season)
            sheet.append(titles)
        wb.remove(wb.get_sheet_by_name("Sheet"))
        wb.save(path + "\\gameData.xlsx")
'''
def addgame2excel(sheetName,gamelist):
    isExists = os.path.exists(path + "\\gameData.xlsx")
    if not isExists:
        print('文件丢失，停止')
        return
    #打开文件
    wb = load_workbook(path + "\\gameData.xlsx")
    wb.get_sheet_by_name(sheetName)
'''




def html2excel():
    sheetNames = gethtmlpath()
    createxls(sheetNames)
    #开始导入数据
    wb = load_workbook(path + "\\gameData.xlsx")#打开文件
    for season in sheetNames:
        seasonSheet = wb.get_sheet_by_name(season)#进入到指定赛季的表
        seasonDir = folderName + "\\"+season
        seasonGames = os.listdir(seasonDir)
        for seasonGame in seasonGames:
            filePath = seasonDir + "\\"+seasonGame
            outputData = analysHtml(filePath)
            for data in outputData:
                seasonSheet.append(data)
    wb.save(path + "\\gameData.xlsx")
